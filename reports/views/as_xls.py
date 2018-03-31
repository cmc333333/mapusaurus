from django.http.response import HttpResponse
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from reports.report_list import report_list
from respondents.models import Institution


def derive_file_name(request):
    lender_id = request.GET.get('lender', '')
    lender = Institution.objects.filter(pk=lender_id).first()
    if lender:
        return slugify(lender.name) + '-report.xlsx'
    else:
        return 'report.xlsx'


def as_xls(request):
    """All reports as a single Excel file with multiple worksheets."""
    workbook = Workbook()
    for report in report_list:
        result = report.fn(request)
        if result.status_code == 400:
            return result
        worksheet = workbook.create_sheet(report.label)
        for x, field_name in enumerate(result.data['fields']):
            worksheet.cell(row=1, column=x + 1, value=field_name)
            for y, row in enumerate(result.data['data']):
                worksheet.cell(row=y + 2, column=x + 1, value=row[field_name])
    del workbook['Sheet']   # remove default, blank sheet
    response = HttpResponse(
        save_virtual_workbook(workbook),
        content_type=('application/vnd.openxmlformats-officedocument.'
                      'spreadsheetml.sheet'),
    )
    filename = derive_file_name(request)
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
